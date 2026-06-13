from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Callable

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont


BASE_DIR = Path(__file__).resolve().parent
RAW_DIR = BASE_DIR / "raw"
EVIDENCE_DIR = BASE_DIR / "evidence"
REPORT_DIR = BASE_DIR / "report"

RAW_RESPONSE_PATH = RAW_DIR / "dbs_response.json"
USAGE_BEFORE_PATH = RAW_DIR / "apollo_usage_before.json"
USAGE_AFTER_PATH = RAW_DIR / "apollo_usage_after.json"
RUN_METADATA_PATH = RAW_DIR / "apollo_run_metadata.json"
REPORT_PATH = REPORT_DIR / "dbs_firmographic.xlsx"
API_DASHBOARD_PATH = EVIDENCE_DIR / "api_dashboard.png"
CREDIT_USAGE_PATH = EVIDENCE_DIR / "credit_usage.png"

APOLLO_BASE_URL = "https://api.apollo.io/api/v1"
ORG_ENRICH_ENDPOINT = "/organizations/enrich"
USAGE_STATS_ENDPOINT = "/usage_stats/api_usage_stats"
ORG_ENRICH_URL = f"{APOLLO_BASE_URL}{ORG_ENRICH_ENDPOINT}"
USAGE_STATS_URL = f"{APOLLO_BASE_URL}{USAGE_STATS_ENDPOINT}"

APOLLO_ORG_DOC = "https://docs.apollo.io/reference/organization-enrichment"
APOLLO_USAGE_DOC = "https://docs.apollo.io/reference/view-api-usage-stats"
APOLLO_PRICING_DOC = "https://docs.apollo.io/docs/api-pricing"

NOT_AVAILABLE = "Not Available"
NOT_VERIFIED = "Not Verified"
SOURCE_API = "Apollo Organization Enrichment"


@dataclass
class FieldResult:
    field: str
    value: str
    source_api: str
    source_field: str
    status: str
    notes: str


def load_dotenv() -> None:
    env_path = BASE_DIR / ".env"
    if not env_path.exists():
        return
    for raw_line in env_path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def ensure_dirs() -> None:
    RAW_DIR.mkdir(exist_ok=True)
    EVIDENCE_DIR.mkdir(exist_ok=True)
    REPORT_DIR.mkdir(exist_ok=True)


def now_iso() -> str:
    return datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")


def save_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def safe_headers(headers: requests.structures.CaseInsensitiveDict[str]) -> dict[str, str]:
    keep_patterns = (
        "content-type",
        "date",
        "retry-after",
        "x-ratelimit",
        "x-rate-limit",
        "rate-limit",
        "ratelimit",
    )
    safe: dict[str, str] = {}
    for key, value in headers.items():
        lowered = key.lower()
        if any(pattern in lowered for pattern in keep_patterns):
            safe[key] = value
    return safe


def auth_headers(api_key: str, mode: str) -> dict[str, str]:
    headers = {
        "Accept": "application/json",
        "Content-Type": "application/json",
        "Cache-Control": "no-cache",
    }
    if mode == "bearer":
        headers["Authorization"] = f"Bearer {api_key}"
    else:
        headers["X-Api-Key"] = api_key
    return headers


def apollo_request(
    method: str,
    url: str,
    api_key: str,
    *,
    params: dict[str, Any] | None = None,
    body: dict[str, Any] | None = None,
    timeout: int = 60,
) -> dict[str, Any]:
    preferred_mode = os.getenv("APOLLO_AUTH_MODE", "").strip().lower()
    auth_modes = [preferred_mode] if preferred_mode in {"x-api-key", "bearer"} else ["x-api-key", "bearer"]
    last_result: dict[str, Any] | None = None

    for mode in auth_modes:
        start = time.perf_counter()
        response = requests.request(
            method,
            url,
            headers=auth_headers(api_key, mode),
            params=params,
            json=body,
            timeout=timeout,
        )
        elapsed_ms = round((time.perf_counter() - start) * 1000, 2)

        try:
            parsed = response.json()
            text = None
        except ValueError:
            parsed = None
            text = response.text

        result = {
            "ok": 200 <= response.status_code < 300,
            "status_code": response.status_code,
            "auth_mode": mode,
            "elapsed_ms": elapsed_ms,
            "headers": safe_headers(response.headers),
            "json": parsed,
            "text": text,
        }
        last_result = result

        if response.status_code not in {401}:
            return result

    return last_result or {
        "ok": False,
        "status_code": None,
        "auth_mode": NOT_AVAILABLE,
        "elapsed_ms": None,
        "headers": {},
        "json": None,
        "text": "No request attempted",
    }


def fetch_usage_stats(api_key: str | None, label: str, output_path: Path) -> dict[str, Any]:
    if not api_key:
        result = {
            "ok": False,
            "status_code": None,
            "label": label,
            "captured_at": now_iso(),
            "reason": "APOLLO_MASTER_API_KEY or APOLLO_API_KEY was not provided.",
            "json": None,
        }
        save_json(output_path, result)
        return result

    result = apollo_request("POST", USAGE_STATS_URL, api_key, body={})
    result["label"] = label
    result["captured_at"] = now_iso()
    save_json(output_path, result)
    return result


def fetch_dbs_from_apollo(api_key: str) -> dict[str, Any]:
    params = {
        "domain": "dbs.com",
        "name": "DBS Group",
        "website": "https://www.dbs.com",
    }
    result = apollo_request("GET", ORG_ENRICH_URL, api_key, params=params)
    body = result["json"] if result["json"] is not None else {"_non_json_response": result.get("text")}
    save_json(RAW_RESPONSE_PATH, body)
    return {"result": result, "params": params}


def is_missing(value: Any) -> bool:
    return value is None or value == "" or value == [] or value == {}


def get_path(data: Any, path: str) -> Any:
    current = data
    for part in path.split("."):
        if isinstance(current, dict) and part in current:
            current = current[part]
        elif isinstance(current, list) and part.isdigit() and int(part) < len(current):
            current = current[int(part)]
        else:
            return None
    return current


def extract_organization(payload: Any) -> tuple[dict[str, Any] | None, str]:
    if not isinstance(payload, dict):
        return None, ""
    for key in ("organization", "company", "account"):
        value = payload.get(key)
        if isinstance(value, dict):
            return value, key
    organizations = payload.get("organizations")
    if isinstance(organizations, list) and organizations and isinstance(organizations[0], dict):
        return organizations[0], "organizations.0"
    if "name" in payload or "primary_domain" in payload or "website_url" in payload:
        return payload, ""
    return None, ""


def source_label(prefix: str, path: str) -> str:
    return f"{prefix}.{path}" if prefix else path


def format_int(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return f"{int(value):,}"
    return str(value)


def format_money(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return f"${int(value):,}"
    return str(value)


def format_percent(value: Any) -> str:
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, (int, float)):
        return f"{value:.2f}%"
    return str(value)


def format_list(value: Any) -> str:
    if isinstance(value, list):
        return ", ".join(str(item) for item in value)
    return str(value)


def first_available(
    org: dict[str, Any] | None,
    org_prefix: str,
    paths: list[str],
    transform: Callable[[Any], str] = str,
) -> tuple[str, str, Any] | None:
    if org is None:
        return None
    for path in paths:
        value = get_path(org, path)
        if not is_missing(value):
            return transform(value), source_label(org_prefix, path), value
    return None


def join_location(org: dict[str, Any] | None, org_prefix: str) -> tuple[str, str] | None:
    if org is None:
        return None
    raw = first_available(org, org_prefix, ["raw_address"])
    if raw:
        return raw[0], raw[1]

    parts: list[str] = []
    source_parts: list[str] = []
    for path in ("street_address", "city", "state", "country", "postal_code"):
        value = get_path(org, path)
        if not is_missing(value):
            parts.append(str(value))
            source_parts.append(source_label(org_prefix, path))
    if parts:
        return ", ".join(parts), ", ".join(source_parts)
    return None


def build_firmographic_rows(payload: Any) -> tuple[list[FieldResult], dict[str, Any]]:
    org, org_prefix = extract_organization(payload)
    rows: list[FieldResult] = []

    def add_available(
        field: str,
        paths: list[str],
        transform: Callable[[Any], str] = str,
        *,
        notes: str = "",
    ) -> None:
        found = first_available(org, org_prefix, paths, transform)
        if found:
            value, field_path, _ = found
            rows.append(FieldResult(field, value, SOURCE_API, field_path, "Available", notes))
        else:
            rows.append(FieldResult(field, NOT_AVAILABLE, SOURCE_API, NOT_AVAILABLE, NOT_AVAILABLE, "Field absent in Apollo response."))

    legal_name = first_available(org, org_prefix, ["legal_name", "registered_name"])
    if legal_name:
        rows.append(FieldResult("Legal Name", legal_name[0], SOURCE_API, legal_name[1], "Available", ""))
    else:
        display_name = first_available(org, org_prefix, ["name"])
        note = "Apollo returned company name only; no explicit legal-name field was present."
        if display_name:
            note = f"{note} API company name: {display_name[0]}."
        rows.append(FieldResult("Legal Name", NOT_VERIFIED, SOURCE_API, display_name[1] if display_name else NOT_AVAILABLE, NOT_VERIFIED, note))

    add_available("Website Domain", ["primary_domain", "domain", "website_domain"])
    add_available("LinkedIn URL", ["linkedin_url", "linkedin"])
    add_available("Employee Count", ["estimated_num_employees", "employee_count", "employees"], format_int)
    add_available(
        "Headcount Growth (1 Year)",
        [
            "organization_headcount_twelve_month_growth",
            "headcount_growth_12_month",
            "headcount_growth_1_year",
            "employee_growth_12_month",
        ],
        format_percent,
    )
    add_available("Revenue / Revenue Band", ["annual_revenue", "estimated_annual_revenue", "revenue", "revenue_range"], format_money)
    add_available("Industry", ["industry"])

    sub_values: list[str] = []
    sub_sources: list[str] = []
    for path in ("naics_codes", "sic_codes", "naics_code", "sic_code"):
        found = first_available(org, org_prefix, [path], format_list)
        if found:
            sub_values.append(found[0])
            sub_sources.append(found[1])
    if sub_values:
        rows.append(FieldResult("Sub-Industry (NAICS/SIC)", " | ".join(sub_values), SOURCE_API, ", ".join(sub_sources), "Available", ""))
    else:
        rows.append(FieldResult("Sub-Industry (NAICS/SIC)", NOT_AVAILABLE, SOURCE_API, NOT_AVAILABLE, NOT_AVAILABLE, "NAICS/SIC field absent in Apollo response."))

    location = join_location(org, org_prefix)
    if location:
        rows.append(FieldResult("Headquarters Location", location[0], SOURCE_API, location[1], "Available", ""))
    else:
        rows.append(FieldResult("Headquarters Location", NOT_AVAILABLE, SOURCE_API, NOT_AVAILABLE, NOT_AVAILABLE, "Location field absent in Apollo response."))

    exact_location_count = first_available(
        org,
        org_prefix,
        ["number_of_locations", "num_locations", "locations_count", "site_count"],
        format_int,
    )
    if exact_location_count:
        rows.append(FieldResult("Number of Sites/Locations (Singapore + Global)", exact_location_count[0], SOURCE_API, exact_location_count[1], "Available", ""))
    else:
        retail_location_count = first_available(org, org_prefix, ["retail_location_count"], format_int)
        if retail_location_count:
            rows.append(
                FieldResult(
                    "Number of Sites/Locations (Singapore + Global)",
                    NOT_VERIFIED,
                    SOURCE_API,
                    retail_location_count[1],
                    NOT_VERIFIED,
                    f"Apollo returned retail_location_count={retail_location_count[0]}, not a verified Singapore + global total.",
                )
            )
        else:
            rows.append(FieldResult("Number of Sites/Locations (Singapore + Global)", NOT_AVAILABLE, SOURCE_API, NOT_AVAILABLE, NOT_AVAILABLE, "Exact location count absent in Apollo response."))

    add_available("Founded Year", ["founded_year", "founded"], format_int)

    explicit_ownership = first_available(org, org_prefix, ["ownership_type", "company_type", "type"])
    if explicit_ownership:
        rows.append(FieldResult("Ownership Type", explicit_ownership[0], SOURCE_API, explicit_ownership[1], "Available", ""))
    else:
        ticker = first_available(org, org_prefix, ["publicly_traded_symbol", "ticker"])
        exchange = first_available(org, org_prefix, ["publicly_traded_exchange", "exchange"])
        if ticker or exchange:
            source_fields = ", ".join(item[1] for item in (ticker, exchange) if item)
            raw_values = ", ".join(item[0] for item in (ticker, exchange) if item)
            rows.append(FieldResult("Ownership Type", "Public Company", SOURCE_API, source_fields, "Inferred from API", f"Inferred from public listing fields returned by Apollo: {raw_values}."))
        else:
            rows.append(FieldResult("Ownership Type", NOT_AVAILABLE, SOURCE_API, NOT_AVAILABLE, NOT_AVAILABLE, "Ownership/listing field absent in Apollo response."))

    add_available("Total Funding Raised", ["total_funding", "funding_total", "total_funding_raised"], format_money)

    context = {"organization_found": org is not None, "organization_prefix": org_prefix}
    return rows, context


def flatten_numbers(data: Any, prefix: str = "") -> dict[str, float]:
    values: dict[str, float] = {}
    if isinstance(data, dict):
        for key, value in data.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            values.update(flatten_numbers(value, next_prefix))
    elif isinstance(data, list):
        for index, value in enumerate(data):
            next_prefix = f"{prefix}.{index}" if prefix else str(index)
            values.update(flatten_numbers(value, next_prefix))
    elif isinstance(data, (int, float)) and not isinstance(data, bool):
        values[prefix] = float(data)
    return values


def find_endpoint_nodes(data: Any, prefix: str = "") -> list[tuple[str, Any]]:
    nodes: list[tuple[str, Any]] = []
    if isinstance(data, dict):
        for key, value in data.items():
            next_prefix = f"{prefix}.{key}" if prefix else str(key)
            lowered = str(key).lower()
            if "organization" in lowered and "enrich" in lowered:
                nodes.append((next_prefix, value))
            nodes.extend(find_endpoint_nodes(value, next_prefix))
    elif isinstance(data, list):
        for index, value in enumerate(data):
            nodes.extend(find_endpoint_nodes(value, f"{prefix}.{index}" if prefix else str(index)))
    return nodes


def usage_json(result: dict[str, Any] | None) -> Any:
    if not result:
        return None
    return result.get("json")


def usage_status(result: dict[str, Any] | None) -> str:
    if not result:
        return NOT_AVAILABLE
    code = result.get("status_code")
    if result.get("ok"):
        return f"Available (HTTP {code})"
    if code == 403:
        return "Not Available (HTTP 403 - master API key required)"
    if code == 401:
        return "Not Available (HTTP 401 - check API key)"
    reason = result.get("reason")
    if reason:
        return reason
    return f"Not Available (HTTP {code})" if code else NOT_AVAILABLE


def summarize_usage_delta(before: dict[str, Any] | None, after: dict[str, Any] | None) -> tuple[str, list[tuple[str, str, str, str]]]:
    before_data = usage_json(before)
    after_data = usage_json(after)
    if not isinstance(before_data, (dict, list)) or not isinstance(after_data, (dict, list)):
        return NOT_VERIFIED, []

    before_numbers = flatten_numbers(before_data)
    after_numbers = flatten_numbers(after_data)
    rows: list[tuple[str, str, str, str]] = []
    for path, after_value in after_numbers.items():
        if path not in before_numbers:
            continue
        before_value = before_numbers[path]
        delta = after_value - before_value
        lowered = path.lower()
        if delta == 0:
            continue
        if any(token in lowered for token in ("credit", "consumed", "used", "usage", "organization", "enrich")):
            rows.append((path, fmt_number(before_value), fmt_number(after_value), fmt_signed_number(delta)))

    if not rows:
        return NOT_VERIFIED, []

    preferred = [row for row in rows if "organization" in row[0].lower() and "enrich" in row[0].lower()]
    summary_rows = preferred or rows
    first = summary_rows[0]
    return f"{first[1]} -> {first[2]} ({first[3]}) at {first[0]}", rows[:20]


def fmt_number(value: float) -> str:
    if value.is_integer():
        return f"{int(value):,}"
    return f"{value:,.2f}"


def fmt_signed_number(value: float) -> str:
    sign = "+" if value >= 0 else ""
    return f"{sign}{fmt_number(value)}"


def summarize_rate_limit(org_result: dict[str, Any] | None, usage_after: dict[str, Any] | None) -> str:
    usage_data = usage_json(usage_after)
    endpoint_nodes = find_endpoint_nodes(usage_data)
    for path, node in endpoint_nodes:
        if isinstance(node, dict):
            parts = []
            for window in ("minute", "hour", "day"):
                window_data = node.get(window)
                if isinstance(window_data, dict):
                    limit = window_data.get("limit")
                    consumed = window_data.get("consumed")
                    left_over = window_data.get("left_over")
                    bits = []
                    if limit is not None:
                        bits.append(f"limit {limit}")
                    if consumed is not None:
                        bits.append(f"consumed {consumed}")
                    if left_over is not None:
                        bits.append(f"left {left_over}")
                    if bits:
                        parts.append(f"{window}: " + ", ".join(bits))
            if parts:
                return f"{path} - " + "; ".join(parts)

    headers = (org_result or {}).get("headers", {})
    rate_headers = [f"{key}: {value}" for key, value in headers.items() if "rate" in key.lower()]
    if rate_headers:
        return "; ".join(rate_headers)
    return NOT_AVAILABLE


def format_status_code(result: dict[str, Any] | None) -> str:
    if not result:
        return NOT_AVAILABLE
    code = result.get("status_code")
    ok = result.get("ok")
    if code is None:
        return NOT_AVAILABLE
    return f"HTTP {code} ({'Success' if ok else 'Failed'})"


def build_validation_rows(
    field_rows: list[FieldResult],
    context: dict[str, Any],
    org_result: dict[str, Any] | None,
    usage_before: dict[str, Any] | None,
    usage_after: dict[str, Any] | None,
    usage_summary: str,
    usage_delta_rows: list[tuple[str, str, str, str]],
) -> list[tuple[str, str]]:
    total_fields = len(field_rows)
    complete_fields = sum(1 for row in field_rows if row.status in {"Available", "Inferred from API"})
    completeness = round((complete_fields / total_fields) * 100, 2) if total_fields else 0
    api_success = bool(org_result and org_result.get("ok") and context.get("organization_found"))
    raw_saved = RAW_RESPONSE_PATH.exists()
    latency = org_result.get("elapsed_ms") if org_result else None
    credits_used = usage_summary
    if usage_delta_rows:
        credits_used = usage_summary

    return [
        ("Tool Name", "Apollo Organization Enrichment API"),
        ("Category", "Firmographic"),
        ("Endpoint Used", f"GET {ORG_ENRICH_URL}"),
        ("API Available", "Y" if api_success else "N"),
        ("Authentication Type", "API Key"),
        ("Free Credits", NOT_VERIFIED),
        ("Credits Used", credits_used),
        ("Credits Before", usage_delta_rows[0][1] if usage_delta_rows else NOT_VERIFIED),
        ("Credits After", usage_delta_rows[0][2] if usage_delta_rows else NOT_VERIFIED),
        ("Rate Limit", summarize_rate_limit(org_result, usage_after)),
        ("Companies Processed", "1"),
        ("Coverage (%)", f"{completeness:.2f}%"),
        ("Success Rate (%)", "100%" if api_success else "0%"),
        ("Error Rate (%)", "0%" if api_success else "100%"),
        ("Average Latency", f"{latency} ms" if latency is not None else NOT_AVAILABLE),
        ("Gated Fields", NOT_VERIFIED),
        ("Free Tier Limitation", "Not Verified - account plan and free-tier limits are not returned in the organization response."),
        ("Paid Plan Cost", "Not Verified - plan-specific pricing was not returned by the API response."),
        ("Paid Tier Benefits", "Not Verified - Apollo docs state advanced API access depends on the account plan."),
        ("Ease of Integration (1-5)", "4" if org_result and org_result.get("ok") else NOT_VERIFIED),
        ("Documentation Quality (1-5)", "4"),
        ("Evidence Link", f"{APOLLO_ORG_DOC}; {APOLLO_USAGE_DOC}; local: raw/dbs_response.json"),
        ("Overall API Score", f"{round((completeness / 20), 1)}/5" if api_success else NOT_VERIFIED),
        ("Status", "Completed" if api_success else "Failed / Not Completed"),
        ("Remarks", build_remarks(org_result, usage_before, usage_after, context)),
        ("Data Completeness (%)", f"{completeness:.2f}%"),
        ("Records Retrieved", "1" if context.get("organization_found") else "0"),
        ("Raw Export Saved", "Y" if raw_saved else "N"),
    ]


def build_remarks(
    org_result: dict[str, Any] | None,
    usage_before: dict[str, Any] | None,
    usage_after: dict[str, Any] | None,
    context: dict[str, Any],
) -> str:
    parts: list[str] = []
    if org_result:
        parts.append(format_status_code(org_result))
    if not context.get("organization_found"):
        parts.append("No organization object found in Apollo response.")
    before_status = usage_status(usage_before)
    after_status = usage_status(usage_after)
    parts.append(f"Usage before: {before_status}")
    parts.append(f"Usage after: {after_status}")
    return " | ".join(parts)


def style_sheet(ws, freeze: str = "A1") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="D8DEE9")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_length = 12
        for cell in ws[letter]:
            if cell.value is None:
                continue
            max_length = max(max_length, min(len(str(cell.value)) + 2, 60))
        ws.column_dimensions[letter].width = max_length


def write_table(ws, start_row: int, headers: list[str], rows: list[list[str]]) -> int:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col, header)
        cell.fill = header_fill
        cell.font = header_font
    for row_index, row in enumerate(rows, start=start_row + 1):
        for col, value in enumerate(row, start=1):
            cell = ws.cell(row_index, col, value)
            if value == NOT_AVAILABLE:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            elif value == NOT_VERIFIED:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            elif str(value).startswith("HTTP 4") or str(value).startswith("HTTP 5"):
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"
    return start_row + len(rows) + 2


def create_workbook(
    field_rows: list[FieldResult],
    validation_rows: list[tuple[str, str]],
    usage_delta_rows: list[tuple[str, str, str, str]],
    metadata: dict[str, Any],
) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    title_fill = PatternFill("solid", fgColor="17365D")
    title_font = Font(color="FFFFFF", bold=True, size=14)

    ws = wb.create_sheet("Firmographic Report")
    ws.merge_cells("A1:F1")
    ws["A1"] = "DBS Group - Apollo Firmographic Sample Report"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    ws["A2"] = f"Generated: {metadata.get('generated_at', now_iso())}"
    rows = [[r.field, r.value, r.source_api, r.source_field, r.status, r.notes] for r in field_rows]
    write_table(ws, 4, ["Field", "Value", "Source API", "Source Field", "Status", "Notes"], rows)
    style_sheet(ws, "A5")

    ws = wb.create_sheet("Missing Data Report")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Missing / Unverified Data"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    missing_rows = []
    for row in field_rows:
        if row.status not in {"Available", "Inferred from API"}:
            missing_rows.append([row.field, row.value, row.source_field, row.notes])
    if not missing_rows:
        missing_rows = [["None", "All required fields available or inferred from API", "", ""]]
    write_table(ws, 3, ["Field", "Value", "Source Field", "Reason"], missing_rows)
    style_sheet(ws, "A4")

    ws = wb.create_sheet("API Validation")
    ws.merge_cells("A1:B1")
    ws["A1"] = "Apollo API Validation Report"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    write_table(ws, 3, ["Field", "Value"], [[key, value] for key, value in validation_rows])
    style_sheet(ws, "A4")

    ws = wb.create_sheet("Run Evidence")
    ws.merge_cells("A1:D1")
    ws["A1"] = "Credit / Usage Evidence"
    ws["A1"].fill = title_fill
    ws["A1"].font = title_font
    next_row = write_table(
        ws,
        3,
        ["Metric", "Value"],
        [
            ["Raw API Response", str(RAW_RESPONSE_PATH.relative_to(BASE_DIR))],
            ["Usage Before Snapshot", str(USAGE_BEFORE_PATH.relative_to(BASE_DIR))],
            ["Usage After Snapshot", str(USAGE_AFTER_PATH.relative_to(BASE_DIR))],
            ["Run Metadata", str(RUN_METADATA_PATH.relative_to(BASE_DIR))],
            ["Apollo Organization Docs", APOLLO_ORG_DOC],
            ["Apollo Usage Docs", APOLLO_USAGE_DOC],
            ["Apollo API Pricing Docs", APOLLO_PRICING_DOC],
        ],
    )
    if usage_delta_rows:
        write_table(ws, next_row, ["Usage Field", "Before", "After", "Delta"], [list(row) for row in usage_delta_rows])
    else:
        write_table(ws, next_row, ["Usage Field", "Before", "After", "Delta"], [[NOT_VERIFIED, NOT_VERIFIED, NOT_VERIFIED, NOT_VERIFIED]])
    style_sheet(ws, "A4")

    wb.save(REPORT_PATH)


def font(name: str, size: int) -> ImageFont.FreeTypeFont | ImageFont.ImageFont:
    candidates = [
        Path("C:/Windows/Fonts") / name,
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/calibri.ttf"),
    ]
    for candidate in candidates:
        if candidate.exists():
            return ImageFont.truetype(str(candidate), size=size)
    return ImageFont.load_default()


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font_obj: ImageFont.ImageFont, max_width: int) -> list[str]:
    words = str(text).split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if draw.textbbox((0, 0), candidate, font=font_obj)[2] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def create_evidence_png(path: Path, title: str, subtitle: str, rows: list[tuple[str, str]]) -> None:
    width = 1500
    margin = 50
    row_padding = 14
    title_font = font("arialbd.ttf", 34)
    subtitle_font = font("arial.ttf", 20)
    header_font = font("arialbd.ttf", 22)
    body_font = font("arial.ttf", 20)
    temp = Image.new("RGB", (width, 200), "white")
    temp_draw = ImageDraw.Draw(temp)
    value_width = width - (margin * 2) - 390
    row_heights: list[int] = []
    for key, value in rows:
        key_lines = wrap_text(temp_draw, key, body_font, 330)
        value_lines = wrap_text(temp_draw, value, body_font, value_width - 40)
        row_heights.append(max(len(key_lines), len(value_lines)) * 28 + row_padding * 2)
    height = 170 + 58 + sum(row_heights) + margin

    image = Image.new("RGB", (width, height), "white")
    draw = ImageDraw.Draw(image)
    draw.rectangle((0, 0, width, 120), fill="#17365D")
    draw.text((margin, 28), title, font=title_font, fill="white")
    draw.text((margin, 130), subtitle, font=subtitle_font, fill="#333333")

    y = 180
    draw.rectangle((margin, y, width - margin, y + 58), fill="#1F4E78")
    draw.text((margin + 18, y + 16), "Metric", font=header_font, fill="white")
    draw.text((margin + 390, y + 16), "Value", font=header_font, fill="white")
    y += 58

    for index, ((key, value), row_height) in enumerate(zip(rows, row_heights)):
        fill = "#F7F9FC" if index % 2 == 0 else "white"
        if value == NOT_AVAILABLE:
            fill = "#FCE4D6"
        elif value == NOT_VERIFIED or value.startswith("Not Verified"):
            fill = "#FFF2CC"
        draw.rectangle((margin, y, width - margin, y + row_height), fill=fill, outline="#D8DEE9")
        draw.line((margin + 370, y, margin + 370, y + row_height), fill="#D8DEE9", width=1)

        key_lines = wrap_text(draw, key, body_font, 330)
        value_lines = wrap_text(draw, value, body_font, value_width - 40)
        text_y = y + row_padding
        for line in key_lines:
            draw.text((margin + 18, text_y), line, font=body_font, fill="#1B1B1B")
            text_y += 28
        text_y = y + row_padding
        for line in value_lines:
            draw.text((margin + 390, text_y), line, font=body_font, fill="#1B1B1B")
            text_y += 28
        y += row_height

    image.save(path)


def create_evidence_images(
    validation_rows: list[tuple[str, str]],
    usage_delta_rows: list[tuple[str, str, str, str]],
    org_result: dict[str, Any] | None,
    usage_before: dict[str, Any] | None,
    usage_after: dict[str, Any] | None,
) -> None:
    lookup = dict(validation_rows)
    dashboard_rows = [
        ("Tool", lookup.get("Tool Name", NOT_AVAILABLE)),
        ("Endpoint", lookup.get("Endpoint Used", NOT_AVAILABLE)),
        ("Status", lookup.get("Status", NOT_AVAILABLE)),
        ("API Response", format_status_code(org_result)),
        ("Average Latency", lookup.get("Average Latency", NOT_AVAILABLE)),
        ("Rate Limit", lookup.get("Rate Limit", NOT_AVAILABLE)),
        ("Usage Before Snapshot", usage_status(usage_before)),
        ("Usage After Snapshot", usage_status(usage_after)),
        ("Raw Export Saved", lookup.get("Raw Export Saved", NOT_AVAILABLE)),
    ]
    create_evidence_png(API_DASHBOARD_PATH, "Apollo API Dashboard Evidence", f"Generated: {now_iso()}", dashboard_rows)

    credit_rows = [
        ("Credits Before", lookup.get("Credits Before", NOT_VERIFIED)),
        ("Credits After", lookup.get("Credits After", NOT_VERIFIED)),
        ("Credits Used", lookup.get("Credits Used", NOT_VERIFIED)),
        ("Token Consumption", "Not Applicable - Apollo uses API credits/usage, not LLM tokens."),
        ("Usage Snapshot File", str(USAGE_AFTER_PATH.relative_to(BASE_DIR))),
    ]
    for field, before, after, delta in usage_delta_rows[:6]:
        credit_rows.append((f"Usage Delta - {field}", f"{before} -> {after} ({delta})"))
    create_evidence_png(CREDIT_USAGE_PATH, "Apollo Credit Usage Evidence", f"Generated: {now_iso()}", credit_rows)


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate DBS Group firmographic validation report from Apollo Organization Enrichment.")
    parser.add_argument("--from-raw", action="store_true", help="Build the report from raw/dbs_response.json without calling Apollo.")
    args = parser.parse_args()

    load_dotenv()
    ensure_dirs()

    api_key = os.getenv("APOLLO_API_KEY") or os.getenv("APOLLO_MASTER_API_KEY")
    usage_key = os.getenv("APOLLO_MASTER_API_KEY") or api_key

    metadata: dict[str, Any] = {
        "generated_at": now_iso(),
        "endpoint": ORG_ENRICH_URL,
        "usage_endpoint": USAGE_STATS_URL,
        "company": "DBS Group",
        "raw_response_path": str(RAW_RESPONSE_PATH.relative_to(BASE_DIR)),
    }

    usage_before: dict[str, Any] | None = None
    usage_after: dict[str, Any] | None = None
    org_result: dict[str, Any] | None = None

    if args.from_raw:
        if not RAW_RESPONSE_PATH.exists():
            print("raw/dbs_response.json is missing. Run with APOLLO_API_KEY set first.", file=sys.stderr)
            return 2
        payload = load_json(RAW_RESPONSE_PATH)
        if USAGE_BEFORE_PATH.exists():
            usage_before = load_json(USAGE_BEFORE_PATH)
        if USAGE_AFTER_PATH.exists():
            usage_after = load_json(USAGE_AFTER_PATH)
        if RUN_METADATA_PATH.exists():
            previous_metadata = load_json(RUN_METADATA_PATH)
            org_result = previous_metadata.get("organization_request", {}).get("result")
    else:
        if not api_key:
            print("Missing Apollo API key. Set APOLLO_API_KEY or APOLLO_MASTER_API_KEY, then rerun this script.", file=sys.stderr)
            return 2
        usage_before = fetch_usage_stats(usage_key, "before", USAGE_BEFORE_PATH)
        fetched = fetch_dbs_from_apollo(api_key)
        org_result = fetched["result"]
        metadata["request_params"] = fetched["params"]
        metadata["organization_request"] = fetched
        usage_after = fetch_usage_stats(usage_key, "after", USAGE_AFTER_PATH)
        metadata["usage_before_status"] = usage_status(usage_before)
        metadata["usage_after_status"] = usage_status(usage_after)
        save_json(RUN_METADATA_PATH, metadata)
        payload = load_json(RAW_RESPONSE_PATH)

    field_rows, context = build_firmographic_rows(payload)
    usage_summary, usage_delta_rows = summarize_usage_delta(usage_before, usage_after)
    validation_rows = build_validation_rows(field_rows, context, org_result, usage_before, usage_after, usage_summary, usage_delta_rows)

    create_workbook(field_rows, validation_rows, usage_delta_rows, metadata)
    create_evidence_images(validation_rows, usage_delta_rows, org_result, usage_before, usage_after)

    print(f"Saved raw response: {RAW_RESPONSE_PATH}")
    print(f"Saved report: {REPORT_PATH}")
    print(f"Saved evidence: {API_DASHBOARD_PATH}")
    print(f"Saved evidence: {CREDIT_USAGE_PATH}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

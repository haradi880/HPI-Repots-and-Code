from __future__ import annotations

import csv
import json
import os
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
ROOT_DIR = BASE_DIR.parent
REPORTS_DIR = BASE_DIR / "reports"

NOT_AVAILABLE = "Not Available"

CATEGORIES = {
    "Firmographic": ROOT_DIR / "Firmographic" / "reports" / "full_10_company_report",
    "Technographic": ROOT_DIR / "Technographic" / "reports",
    "Jobs / Hiring": ROOT_DIR / "JobsHiring" / "reports",
    "Contact-Level": ROOT_DIR / "ContactLevel" / "reports",
    "News / Key Announcements": ROOT_DIR / "NewsAnnouncements" / "reports",
}

TRACE_FILES = {
    "Firmographic": [
        ROOT_DIR / "Firmographic" / "reports" / "full_10_company_report" / "api_trace_full_report.csv",
        ROOT_DIR / "Firmographic" / "reports" / "full_10_company_report" / "coresignal_api_trace_report.csv",
    ],
    "Technographic": [ROOT_DIR / "Technographic" / "reports" / "api_tracing_report.csv"],
    "Jobs / Hiring": [ROOT_DIR / "JobsHiring" / "reports" / "api_tracing_report.csv"],
    "Contact-Level": [ROOT_DIR / "ContactLevel" / "reports" / "api_tracing_report.csv"],
    "News / Key Announcements": [ROOT_DIR / "NewsAnnouncements" / "reports" / "api_tracing_report.csv"],
}

COMPARISON_FILES = {
    "Firmographic": ROOT_DIR / "Firmographic" / "reports" / "full_10_company_report" / "api_comparison_report.csv",
    "Technographic": ROOT_DIR / "Technographic" / "reports" / "api_comparison_report.csv",
    "Jobs / Hiring": ROOT_DIR / "JobsHiring" / "reports" / "api_comparison_report.csv",
    "Contact-Level": ROOT_DIR / "ContactLevel" / "reports" / "api_comparison_report.csv",
    "News / Key Announcements": ROOT_DIR / "NewsAnnouncements" / "reports" / "api_comparison_report.csv",
}

CALL_LOG_FILES = {
    "Technographic": ROOT_DIR / "Technographic" / "reports" / "api_call_log.csv",
    "Jobs / Hiring": ROOT_DIR / "JobsHiring" / "reports" / "api_call_log.csv",
    "Contact-Level": ROOT_DIR / "ContactLevel" / "reports" / "api_call_log.csv",
    "News / Key Announcements": ROOT_DIR / "NewsAnnouncements" / "reports" / "api_call_log.csv",
}

FINAL_REPORT_PATTERNS = {
    "Firmographic": [
        "hpi_10_company_firmographic_api_comparison_apollo_coresignal.xlsx",
        "hpi_10_company_firmographic_api_comparison_apollo_coresignal.docx",
        "hpi_10_company_coresignal_firmographic_api_report.xlsx",
        "hpi_10_company_coresignal_firmographic_api_report.docx",
    ],
    "Technographic": ["hpi_technographic_api_comparison_20260613_184706.xlsx", "hpi_technographic_api_comparison_20260613_184706.docx"],
    "Jobs / Hiring": ["hpi_jobs_hiring_api_evaluation_20260613_181454.xlsx", "hpi_jobs_hiring_api_evaluation_20260613_181454.docx"],
    "Contact-Level": ["hpi_contact_level_api_evaluation_20260613_190106.xlsx", "hpi_contact_level_api_evaluation_20260613_190106.docx"],
    "News / Key Announcements": ["hpi_news_announcements_api_evaluation_20260614_124248.xlsx", "hpi_news_announcements_api_evaluation_20260614_124248.docx"],
}

KEY_ALIASES = {
    "Apollo": ["APOLLO_API_KEY", "APOLLO_MASTER_API_KEY"],
    "Coresignal": ["CORESIGNAL_API_KEY"],
    "TheirStack": ["THEIRSTACK_API_KEY"],
    "PredictLeads": ["PREDICTLEADS_API_KEY"],
    "FullEnrich": ["FULLENRICH_API_KEY", "FULLENRICH_API_TOKEN"],
    "Prospeo": ["PROSPEO_API_KEY"],
    "SignalHire": ["SIGNALHIRE_API_KEY"],
    "People Data Labs": ["PEOPLE_DATA_LABS_API_KEY"],
    "Exa": ["EXA_API_KEY"],
    "Tavily": ["TAVILY_API_KEY"],
    "NewsAPI": ["NEWSAPI_API_KEY"],
}

MANUAL_KEY_STATUS = {
    "Google News RSS": ("Not required", "No API key"),
    "GDELT": ("Not required", "No API key"),
    "LinkUp / Aura": ("Sample on request", "No key used"),
}


def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")


def rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT_DIR)).replace("\\", "/")
    except ValueError:
        return str(path).replace("\\", "/")


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_env_values() -> dict[str, str]:
    values: dict[str, str] = {}
    for env_path in ROOT_DIR.glob("*/.env"):
        for raw in env_path.read_text(encoding="utf-8-sig").splitlines():
            line = raw.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            values.setdefault(key.strip().lstrip("\ufeff"), value.strip().strip('"').strip("'"))
    return values


def mask_key(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    if len(value) <= 10:
        return value[:2] + "***" + value[-2:]
    return f"{value[:4]}...{value[-4:]}"


def tool_family(tool_name: str) -> str:
    text = tool_name.lower()
    if "apollo" in text:
        return "Apollo"
    if "coresignal" in text:
        return "Coresignal"
    if "theirstack" in text:
        return "TheirStack"
    if "predictleads" in text:
        return "PredictLeads"
    if "linkup" in text or "aura" in text:
        return "LinkUp / Aura"
    if "fullenrich" in text:
        return "FullEnrich"
    if "prospeo" in text:
        return "Prospeo"
    if "signalhire" in text:
        return "SignalHire"
    if "people data labs" in text:
        return "People Data Labs"
    if "exa" in text:
        return "Exa"
    if "tavily" in text:
        return "Tavily"
    if "google news" in text:
        return "Google News RSS"
    if "newsapi" in text:
        return "NewsAPI"
    if "gdelt" in text:
        return "GDELT"
    return tool_name


def api_key_status(tool_name: str, env_values: dict[str, str]) -> tuple[str, str]:
    family = tool_family(tool_name)
    if family in MANUAL_KEY_STATUS:
        return MANUAL_KEY_STATUS[family]
    for key_name in KEY_ALIASES.get(family, []):
        value = env_values.get(key_name) or os.environ.get(key_name, "")
        if value:
            return "Available", mask_key(value)
    return "Missing / Not used", ""


def clean_percent(value: str) -> str:
    return (value or "").replace("%", "")


def collect_trace_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    env_values = load_env_values()
    for category, paths in TRACE_FILES.items():
        seen_tools: set[str] = set()
        for path in paths:
            for row in read_csv(path):
                tool = row.get("Tool Name") or row.get("api_name") or row.get("tool") or ""
                if not tool or tool in seen_tools:
                    continue
                seen_tools.add(tool)
                key_status, masked_key = api_key_status(tool, env_values)
                rows.append(
                    {
                        "Class / Category": category,
                        "Tool": tool,
                        "API Status": row.get("Status") or row.get("success_failure") or NOT_AVAILABLE,
                        "API Used / Endpoint": row.get("Evidence Link") or row.get("endpoint_used") or NOT_AVAILABLE,
                        "API Key Status": key_status,
                        "API Key (masked)": masked_key,
                        "Reason / Remarks": row.get("Remarks") or row.get("Gated Fields") or row.get("error_message") or "",
                        "Credits / Tokens Used": row.get("Credits / Tokens Used") or row.get("credits_consumed") or NOT_AVAILABLE,
                        "Rate Limit": row.get("Rate Limit") or row.get("rate_limits") or NOT_AVAILABLE,
                        "Average Latency": row.get("Average Latency") or row.get("latency_ms") or NOT_AVAILABLE,
                        "Companies Processed": row.get("Companies Processed") or row.get("Companies Processed ") or NOT_AVAILABLE,
                        "Success Rate (%)": clean_percent(row.get("Success Rate (%)") or row.get("Coverage (%)") or ""),
                        "Data Completeness (%)": clean_percent(row.get("Data Completeness (%)") or row.get("field_completeness_percent") or ""),
                        "Records Retrieved": row.get("Records Retrieved") or row.get("records_retrieved") or NOT_AVAILABLE,
                        "Raw Export Saved": row.get("Raw Export Saved (Y/N)") or NOT_AVAILABLE,
                    }
                )
    return rows


def collect_comparison_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    env_values = load_env_values()
    for category, path in COMPARISON_FILES.items():
        for row in read_csv(path):
            tool = row.get("API Name", "")
            key_status, masked_key = api_key_status(tool, env_values)
            rows.append(
                {
                    "Class / Category": category,
                    "Tool": tool,
                    "API Status": row.get("Status (Success/Fail)", NOT_AVAILABLE),
                    "API Used / Endpoint": row.get("Endpoint Used", NOT_AVAILABLE),
                    "API Key Status": key_status,
                    "API Key (masked)": masked_key,
                    "Fields Returned": row.get("Fields Returned", NOT_AVAILABLE),
                    "Reason / Notes": row.get("Notes", ""),
                    "Free-Tier Limitations": row.get("Free-Tier Limitations", ""),
                    "Paid-Tier Benefits": row.get("Paid-Tier Benefits", ""),
                }
            )
    return rows


def collect_call_summary_rows() -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for category, path in CALL_LOG_FILES.items():
        grouped: dict[str, list[dict[str, str]]] = {}
        for row in read_csv(path):
            grouped.setdefault(row.get("api_name", ""), []).append(row)
        for tool, calls in sorted(grouped.items()):
            latencies = []
            records = 0
            statuses = Counter()
            credits = []
            for call in calls:
                statuses[call.get("status", "")] += 1
                if call.get("latency_ms", "").replace(".", "", 1).isdigit():
                    latencies.append(float(call["latency_ms"]))
                if call.get("records_retrieved", "").isdigit():
                    records += int(call["records_retrieved"])
                if call.get("credits_used"):
                    credits.append(call["credits_used"])
            rows.append(
                {
                    "Class / Category": category,
                    "Tool": tool,
                    "Calls / Rows": len(calls),
                    "Success Calls": statuses.get("Success", 0),
                    "Fail Calls": statuses.get("Fail", 0),
                    "Not Tested Calls": statuses.get("Not Tested", 0),
                    "Average Latency (ms)": round(sum(latencies) / len(latencies), 2) if latencies else NOT_AVAILABLE,
                    "Records Retrieved": records,
                    "Credits / Tokens Evidence": "; ".join(sorted(set(credits)))[:1200] if credits else NOT_AVAILABLE,
                }
            )
    return rows


def collect_report_inventory() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for category, report_names in FINAL_REPORT_PATTERNS.items():
        base = CATEGORIES[category]
        for name in report_names:
            path = base / name
            rows.append(
                {
                    "Class / Category": category,
                    "Report File": rel(path),
                    "Exists": "Y" if path.exists() else "N",
                    "File Type": path.suffix.lstrip(".").upper(),
                    "Size KB": f"{path.stat().st_size / 1024:.1f}" if path.exists() else "",
                }
            )
    return rows


def overview_rows(trace_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    by_category: dict[str, list[dict[str, str]]] = {}
    for row in trace_rows:
        by_category.setdefault(row["Class / Category"], []).append(row)
    for category in CATEGORIES:
        rows = by_category.get(category, [])
        success = sum(1 for row in rows if "success" in row.get("API Status", "").lower())
        blocked = len(rows) - success
        records = sum(int(row["Records Retrieved"]) for row in rows if str(row.get("Records Retrieved", "")).isdigit())
        out.append(
            {
                "Class / Category": category,
                "Tools Evaluated": str(len(rows)),
                "Successful Tools": str(success),
                "Blocked / Not Tested Tools": str(blocked),
                "Records Retrieved": str(records),
                "Final Status": "Complete",
                "Notes": "Excel-only consolidated summary; detailed raw data stays in each category folder.",
            }
        )
    return out


def write_sheet(ws, rows: list[dict[str, Any]], fields: list[str]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    sub_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    border = Border(left=Side(style="thin", color="D9D9D9"), right=Side(style="thin", color="D9D9D9"), top=Side(style="thin", color="D9D9D9"), bottom=Side(style="thin", color="D9D9D9"))
    ws.append(fields)
    for row in rows:
        ws.append([row.get(field, "") for field in fields])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        if row[0].value in CATEGORIES:
            row[0].fill = sub_fill
    ws.freeze_panes = "A2"
    for column in ws.columns:
        col = get_column_letter(column[0].column)
        width = max(12, min(max(len(str(cell.value or "")) for cell in column) + 2, 55))
        ws.column_dimensions[col].width = width


def build_workbook() -> Path:
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    trace_rows = collect_trace_rows()
    comparison_rows = collect_comparison_rows()
    call_rows = collect_call_summary_rows()
    reports = collect_report_inventory()
    overview = overview_rows(trace_rows)

    path = REPORTS_DIR / f"hpi_final_api_status_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    sheets = {
        "Overview": (
            overview,
            ["Class / Category", "Tools Evaluated", "Successful Tools", "Blocked / Not Tested Tools", "Records Retrieved", "Final Status", "Notes"],
        ),
        "API Status Table": (
            comparison_rows,
            ["Class / Category", "Tool", "API Status", "API Used / Endpoint", "API Key Status", "API Key (masked)", "Fields Returned", "Reason / Notes", "Free-Tier Limitations", "Paid-Tier Benefits"],
        ),
        "API Extraction Metrics": (
            trace_rows,
            ["Class / Category", "Tool", "API Status", "API Used / Endpoint", "API Key Status", "API Key (masked)", "Credits / Tokens Used", "Rate Limit", "Average Latency", "Companies Processed", "Success Rate (%)", "Data Completeness (%)", "Records Retrieved", "Raw Export Saved", "Reason / Remarks"],
        ),
        "Call Log Summary": (
            call_rows,
            ["Class / Category", "Tool", "Calls / Rows", "Success Calls", "Fail Calls", "Not Tested Calls", "Average Latency (ms)", "Records Retrieved", "Credits / Tokens Evidence"],
        ),
        "Final Report Inventory": (
            reports,
            ["Class / Category", "Report File", "Exists", "File Type", "Size KB"],
        ),
    }
    for sheet_name, (rows, fields) in sheets.items():
        write_sheet(wb.create_sheet(sheet_name), rows, fields)
    meta = wb.create_sheet("Read Me")
    write_sheet(
        meta,
        [
            {"Field": "Generated At", "Value": now_iso()},
            {"Field": "Purpose", "Value": "Final Excel-only API status summary across all HPI evaluation classes. This file intentionally does not include company-level raw data."},
            {"Field": "API Keys", "Value": "Keys are masked before commit. Full secrets remain only in ignored local .env files."},
            {"Field": "Included Classes", "Value": "; ".join(CATEGORIES.keys())},
        ],
        ["Field", "Value"],
    )
    wb.save(path)
    return path


if __name__ == "__main__":
    output = build_workbook()
    print(rel(output))
